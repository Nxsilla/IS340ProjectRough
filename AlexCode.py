
#Open the excel file
#from pandas import *
from statistics import *
from textblob import TextBlob
import pandas as pd
from openpyxl import *

 
df = pd.read_excel ('Movie_Ratings.xlsx')
pd.set_option('display.max_columns', None)


#excel file
excelFile=load_workbook("Movie_Ratings.xlsx")
 
#active method 2
sheet1=excelFile.active
 
 
 
"""
 
Using only the first genre listed for each movie line
 
Make a list of all ratings in each genre
 
Find average, max, min, etc. for each genre
 
"""

#reads Movie Name, Tagline, and Genres Label
movieNameLabel=sheet1.cell(column=2, row=1).value
taglineLabel=sheet1.cell(column=4, row=1).value
genreLabel=sheet1.cell(column=5, row=1).value

#prints labels
print(movieNameLabel,",", genreLabel, ",", taglineLabel)


#used to stop in while loop
serial_id=sheet1.cell(column=1, row=1).value

i=2

#ratingList=[]

#program reads "product color", "rating", and "rating text"

while serial_id != None:

    #Get value
    serial_id=sheet1.cell(column=1, row=i).value
    
    movieNameColumn=sheet1.cell(column=2, row=i).value
        
    genreColumn=sheet1.cell(column=5, row=i).value

    taglineColumn=sheet1.cell(column=4, row=i).value


    #Sanitize
    if movieNameColumn == None:
        movieNameColumn = ""
    if genreColumn == None:
        genreColumn = ""
    if taglineColumn == None:
        taglineColumn = ""

    #Print
    formattedColumn = "{:45} {:75} {:.70}"
    print(formattedColumn.format(movieNameColumn, genreColumn, taglineColumn ))

    i=i+1

 
# Construct genre list from incoming data
pd_genre_list = df["Genres"].astype(str).tolist()
"""
#print ("\n".join(pd_genre_list))
 
my_genre_list = []
 
for genre_text in pd_genre_list:
   if " " in genre_text:
       #split it and pick first one
       #XXXXXX
       pass
 
   else:
       my_genre_list.append(genre_text)
 
print(my_genre_list)
 
#eliminate duplicates XXXXXX
 
"""
 
my_genre_list = ['Drama','Western','Action','Horror','Adventure','Thriller','War','Comedy','Crime','Mystry', 'Romance', 'Biography']
 
for genre in my_genre_list:
   print ("\nAnalyzing genre: " + genre)
 
   ratings = []
 
   for idx in range(len(pd_genre_list)):
       if genre in pd_genre_list[idx]:
           ratings.append(df.at[idx,"Rating"])
 
   #print(ratings)
   print ("Movies in genre: ", len(ratings))
   print ("Min rating: ", min(ratings))
   print ("Max rating: ", max(ratings))
   print ("Average rating: ", mean(ratings))
   print ("Standard Deviation: ", stdev(ratings))
